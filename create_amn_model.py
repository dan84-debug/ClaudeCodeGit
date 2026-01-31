from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = 'Model'

# Define styles
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
number_format = '#,##0'
pct_format = '0.0%'

# Light gray fill for driver rows
driver_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

# Column layout:
# A = marker column (x)
# B = labels
# C-F = Annual historical (FY 2021-2024)
# G-L = Annual forecast (FY 2025-2030)
# M-Q = blank separator (5 columns)
# R-U = Quarterly 2024 (Q1-Q4 2024)
# V-Y = Quarterly 2025 (Q1-Q4 2025)
# Z-AC = Quarterly 2026 forecast (Q1-Q4 2026)

# Historical annual data
annual_years_hist = ['FY 2021', 'FY 2022', 'FY 2023', 'FY 2024']
annual_dates_hist = ['12/31/2021', '12/31/2022', '12/31/2023', '12/31/2024']
annual_cols_hist = ['C', 'D', 'E', 'F']

# Forecast annual data
annual_years_fcast = ['FY 2025', 'FY 2026', 'FY 2027', 'FY 2028', 'FY 2029', 'FY 2030']
annual_dates_fcast = ['12/31/2025', '12/31/2026', '12/31/2027', '12/31/2028', '12/31/2029', '12/31/2030']
annual_cols_fcast = ['G', 'H', 'I', 'J', 'K', 'L']

# All annual columns combined
annual_years = annual_years_hist + annual_years_fcast
annual_dates = annual_dates_hist + annual_dates_fcast
annual_cols = annual_cols_hist + annual_cols_fcast

# Quarterly 2024 (historical)
quarterly_periods_2024 = ['Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024']
quarterly_dates_2024 = ['3/31/2024', '6/30/2024', '9/30/2024', '12/31/2024']
quarterly_cols_2024 = ['R', 'S', 'T', 'U']

# Quarterly 2025 (historical/current)
quarterly_periods_2025 = ['Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025']
quarterly_dates_2025 = ['3/31/2025', '6/30/2025', '9/30/2025', '12/31/2025']
quarterly_cols_2025 = ['V', 'W', 'X', 'Y']

# Quarterly 2026 (forecast)
quarterly_periods_2026 = ['Q1 2026', 'Q2 2026', 'Q3 2026', 'Q4 2026']
quarterly_dates_2026 = ['3/31/2026', '6/30/2026', '9/30/2026', '12/31/2026']
quarterly_cols_2026 = ['Z', 'AA', 'AB', 'AC']

# All quarterly combined
quarterly_periods = quarterly_periods_2024 + quarterly_periods_2025 + quarterly_periods_2026
quarterly_dates = quarterly_dates_2024 + quarterly_dates_2025 + quarterly_dates_2026
quarterly_cols = quarterly_cols_2024 + quarterly_cols_2025 + quarterly_cols_2026

# Segment Revenue Data (in thousands) - HISTORICAL
segment_annual = {
    'nurse_allied': [2990100, 3982500, 2624500, 1815700],
    'physician_leadership': [594200, 697900, 669700, 728600],
    'tech_workforce': [399900, 562800, 495000, 439500],
}

# Q1-Q4 2024 segment data
segment_quarterly_2024 = {
    'nurse_allied': [519300, 442400, 399400, 454700],
    'physician_leadership': [188800, 186100, 180600, 173100],
    'tech_workforce': [112800, 112200, 107500, 106900],
}

# Q1-Q4 2025 segment data (placeholder - to be filled with actuals)
segment_quarterly_2025 = {
    'nurse_allied': [0, 0, 0, 0],
    'physician_leadership': [0, 0, 0, 0],
    'tech_workforce': [0, 0, 0, 0],
}

# Prior year quarterly for YoY (Q1-Q4 2023)
segment_quarterly_prior_2023 = {
    'nurse_allied': [756900, 756600, 573400, 537600],
    'physician_leadership': [170900, 171000, 159600, 168200],
    'tech_workforce': [131200, 130800, 120500, 112500],
}

# Income Statement Data (in thousands) - HISTORICAL ANNUAL
annual_data = {
    'revenue': [3984235, 5243242, 3789254, 2983781],
    'cost_of_revenue': [2674634, 3526558, 2539673, 2064405],
    'sga': [730451, 936576, 756238, 632489],
    'da': [101152, 133007, 154914, 167103],
    'goodwill_impairment': [0, 0, 0, 222457],
    'interest_expense': [34077, 40398, 54140, 69901],
    'tax_expense': [116533, 162653, 73610, -25595],
    # Balance Sheet - Assets
    'cash': [180928, 64524, 32935, 10649],
    'accounts_receivable': [789131, 675650, 623488, 437817],
    'subcontractor_recv': [239719, 268726, 117703, 70481],
    'prepaid_other': [139290, 84745, 67559, 75968],
    'restricted_cash': [64482, 61218, 68845, 71840],
    'fixed_assets': [127114, 149276, 191385, 186270],
    'other_assets': [156670, 172016, 236796, 258053],
    'deferred_tax_asset': [0, 0, 0, 25829],
    'goodwill': [892341, 935364, 1111549, 897456],
    'intangibles': [514460, 476832, 474134, 381364],
    # Balance Sheet - Liabilities
    'ap_accrued': [425257, 476452, 343847, 184311],
    'accrued_compensation': [354381, 333244, 278536, 287544],
    'other_current_liab': [189752, 48237, 33738, 73930],
    'revolving_credit': [0, 0, 460000, 210000],
    'notes_payable': [842322, 843505, 844688, 845872],
    'deferred_tax_liab': [47814, 22713, 23350, 0],
    'other_lt_liab': [110353, 120566, 108979, 107450],
}

# Quarterly 2024 Income Statement Data (historical)
quarterly_data_2024 = {
    'revenue': [820878, 740685, 687509, 734709],
    'cost_of_revenue': [563372, 510858, 474454, 515721],
    'sga': [174842, 149044, 149681, 158922],
    'da': [42719, 43101, 41122, 40161],
    'goodwill_impairment': [0, 0, 0, 222457],
    'interest_expense': [16628, 15715, 14444, 23114],
    'tax_expense': [5989, 5730, 819, -38133],
}

# Quarterly 2025 Income Statement Data (placeholder - to be filled with actuals)
quarterly_data_2025 = {
    'revenue': [0, 0, 0, 0],
    'cost_of_revenue': [0, 0, 0, 0],
    'sga': [0, 0, 0, 0],
    'da': [0, 0, 0, 0],
    'goodwill_impairment': [0, 0, 0, 0],
    'interest_expense': [0, 0, 0, 0],
    'tax_expense': [0, 0, 0, 0],
}

# Prior year quarterly revenue for YoY growth (Q1-Q4 2023)
quarterly_revenue_prior_2023 = [1126223, 991299, 853463, 818269]

# Track row references for formulas
row_refs = {}

# Start writing data
row = 1

# Title
ws.cell(row=row, column=2, value='AMN Healthcare Services Inc (AMN) - Financial Model').font = title_font
row += 2

# ==================== SEGMENT REVENUE BREAKDOWN ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='SEGMENT REVENUE BREAKDOWN - In Thousands USD').font = header_font
# Historical annual headers
for i, yr in enumerate(annual_years_hist):
    ws.cell(row=row, column=3+i, value=yr).font = header_font
# Forecast annual headers
for i, yr in enumerate(annual_years_fcast):
    ws.cell(row=row, column=7+i, value=yr).font = header_font
# Quarterly headers
for i, qtr in enumerate(quarterly_periods):
    ws.cell(row=row, column=18+i, value=qtr).font = header_font
row += 1

# Nurse and Allied Solutions
row_refs['seg_nurse'] = row
ws.cell(row=row, column=2, value='Nurse and Allied Solutions')
for i, val in enumerate(segment_annual['nurse_allied']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
# Leave forecast annual columns empty (will be sum of quarters or manual input)
for i, val in enumerate(segment_quarterly_2024['nurse_allied']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(segment_quarterly_2025['nurse_allied']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
# Q1-Q4 2026 left blank for projections
row += 1

# Nurse YoY Growth
row_refs['seg_nurse_yoy'] = row
ws.cell(row=row, column=2, value='  YoY Growth')
# Annual YoY: current / prior - 1
for i, col in enumerate(annual_cols_hist):
    if i == 0:
        ws.cell(row=row, column=3+i, value='')
    else:
        prior_col = annual_cols_hist[i-1]
        formula = f'={col}{row_refs["seg_nurse"]}/{prior_col}{row_refs["seg_nurse"]}-1'
        ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
# Forecast annual YoY
for i, col in enumerate(annual_cols_fcast):
    if i == 0:
        prior_col = annual_cols_hist[-1]  # FY 2024
    else:
        prior_col = annual_cols_fcast[i-1]
    formula = f'=IF({col}{row_refs["seg_nurse"]}>0,{col}{row_refs["seg_nurse"]}/{prior_col}{row_refs["seg_nurse"]}-1,"")'
    ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
# Quarterly YoY (vs prior year same quarter)
for i, val in enumerate(segment_quarterly_prior_2023['nurse_allied']):
    if val > 0:
        formula = f'={quarterly_cols_2024[i]}{row_refs["seg_nurse"]}/{val}-1'
        ws.cell(row=row, column=18+i, value=formula).number_format = pct_format
# Q1-Q4 2025 YoY vs Q1-Q4 2024
for i in range(4):
    formula = f'=IF({quarterly_cols_2025[i]}{row_refs["seg_nurse"]}>0,{quarterly_cols_2025[i]}{row_refs["seg_nurse"]}/{quarterly_cols_2024[i]}{row_refs["seg_nurse"]}-1,"")'
    ws.cell(row=row, column=22+i, value=formula).number_format = pct_format
# Q1-Q4 2026 YoY vs Q1-Q4 2025
for i in range(4):
    formula = f'=IF({quarterly_cols_2026[i]}{row_refs["seg_nurse"]}>0,{quarterly_cols_2026[i]}{row_refs["seg_nurse"]}/{quarterly_cols_2025[i]}{row_refs["seg_nurse"]}-1,"")'
    ws.cell(row=row, column=26+i, value=formula).number_format = pct_format
row += 1

# Physician and Leadership Solutions
row_refs['seg_physician'] = row
ws.cell(row=row, column=2, value='Physician and Leadership Solutions')
for i, val in enumerate(segment_annual['physician_leadership']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
for i, val in enumerate(segment_quarterly_2024['physician_leadership']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(segment_quarterly_2025['physician_leadership']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
row += 1

# Physician YoY Growth
row_refs['seg_physician_yoy'] = row
ws.cell(row=row, column=2, value='  YoY Growth')
for i, col in enumerate(annual_cols_hist):
    if i == 0:
        ws.cell(row=row, column=3+i, value='')
    else:
        prior_col = annual_cols_hist[i-1]
        formula = f'={col}{row_refs["seg_physician"]}/{prior_col}{row_refs["seg_physician"]}-1'
        ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
for i, col in enumerate(annual_cols_fcast):
    if i == 0:
        prior_col = annual_cols_hist[-1]
    else:
        prior_col = annual_cols_fcast[i-1]
    formula = f'=IF({col}{row_refs["seg_physician"]}>0,{col}{row_refs["seg_physician"]}/{prior_col}{row_refs["seg_physician"]}-1,"")'
    ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
for i, val in enumerate(segment_quarterly_prior_2023['physician_leadership']):
    if val > 0:
        formula = f'={quarterly_cols_2024[i]}{row_refs["seg_physician"]}/{val}-1'
        ws.cell(row=row, column=18+i, value=formula).number_format = pct_format
for i in range(4):
    formula = f'=IF({quarterly_cols_2025[i]}{row_refs["seg_physician"]}>0,{quarterly_cols_2025[i]}{row_refs["seg_physician"]}/{quarterly_cols_2024[i]}{row_refs["seg_physician"]}-1,"")'
    ws.cell(row=row, column=22+i, value=formula).number_format = pct_format
for i in range(4):
    formula = f'=IF({quarterly_cols_2026[i]}{row_refs["seg_physician"]}>0,{quarterly_cols_2026[i]}{row_refs["seg_physician"]}/{quarterly_cols_2025[i]}{row_refs["seg_physician"]}-1,"")'
    ws.cell(row=row, column=26+i, value=formula).number_format = pct_format
row += 1

# Technology and Workforce Solutions
row_refs['seg_tech'] = row
ws.cell(row=row, column=2, value='Technology and Workforce Solutions')
for i, val in enumerate(segment_annual['tech_workforce']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
for i, val in enumerate(segment_quarterly_2024['tech_workforce']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(segment_quarterly_2025['tech_workforce']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
row += 1

# Tech YoY Growth
row_refs['seg_tech_yoy'] = row
ws.cell(row=row, column=2, value='  YoY Growth')
for i, col in enumerate(annual_cols_hist):
    if i == 0:
        ws.cell(row=row, column=3+i, value='')
    else:
        prior_col = annual_cols_hist[i-1]
        formula = f'={col}{row_refs["seg_tech"]}/{prior_col}{row_refs["seg_tech"]}-1'
        ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
for i, col in enumerate(annual_cols_fcast):
    if i == 0:
        prior_col = annual_cols_hist[-1]
    else:
        prior_col = annual_cols_fcast[i-1]
    formula = f'=IF({col}{row_refs["seg_tech"]}>0,{col}{row_refs["seg_tech"]}/{prior_col}{row_refs["seg_tech"]}-1,"")'
    ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
for i, val in enumerate(segment_quarterly_prior_2023['tech_workforce']):
    if val > 0:
        formula = f'={quarterly_cols_2024[i]}{row_refs["seg_tech"]}/{val}-1'
        ws.cell(row=row, column=18+i, value=formula).number_format = pct_format
for i in range(4):
    formula = f'=IF({quarterly_cols_2025[i]}{row_refs["seg_tech"]}>0,{quarterly_cols_2025[i]}{row_refs["seg_tech"]}/{quarterly_cols_2024[i]}{row_refs["seg_tech"]}-1,"")'
    ws.cell(row=row, column=22+i, value=formula).number_format = pct_format
for i in range(4):
    formula = f'=IF({quarterly_cols_2026[i]}{row_refs["seg_tech"]}>0,{quarterly_cols_2026[i]}{row_refs["seg_tech"]}/{quarterly_cols_2025[i]}{row_refs["seg_tech"]}-1,"")'
    ws.cell(row=row, column=26+i, value=formula).number_format = pct_format
row += 1

# Total Revenue (formula summing segments)
row_refs['seg_total'] = row
ws.cell(row=row, column=2, value='Total Revenue').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["seg_nurse"]}+{col}{row_refs["seg_physician"]}+{col}{row_refs["seg_tech"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["seg_nurse"]}+{col}{row_refs["seg_physician"]}+{col}{row_refs["seg_tech"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
for i, col in enumerate(quarterly_cols):
    formula = f'={col}{row_refs["seg_nurse"]}+{col}{row_refs["seg_physician"]}+{col}{row_refs["seg_tech"]}'
    ws.cell(row=row, column=18+i, value=formula).number_format = number_format
row += 1

# Total YoY Growth
row_refs['seg_total_yoy'] = row
ws.cell(row=row, column=2, value='  YoY Growth').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    if i == 0:
        ws.cell(row=row, column=3+i, value='')
    else:
        prior_col = annual_cols_hist[i-1]
        formula = f'={col}{row_refs["seg_total"]}/{prior_col}{row_refs["seg_total"]}-1'
        ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
for i, col in enumerate(annual_cols_fcast):
    if i == 0:
        prior_col = annual_cols_hist[-1]
    else:
        prior_col = annual_cols_fcast[i-1]
    formula = f'=IF({col}{row_refs["seg_total"]}>0,{col}{row_refs["seg_total"]}/{prior_col}{row_refs["seg_total"]}-1,"")'
    ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
# Quarterly 2024 YoY
for i, val in enumerate(quarterly_revenue_prior_2023):
    if val > 0:
        formula = f'={quarterly_cols_2024[i]}{row_refs["seg_total"]}/{val}-1'
        ws.cell(row=row, column=18+i, value=formula).number_format = pct_format
# Q1-Q4 2025 YoY
for i in range(4):
    formula = f'=IF({quarterly_cols_2025[i]}{row_refs["seg_total"]}>0,{quarterly_cols_2025[i]}{row_refs["seg_total"]}/{quarterly_cols_2024[i]}{row_refs["seg_total"]}-1,"")'
    ws.cell(row=row, column=22+i, value=formula).number_format = pct_format
# Q1-Q4 2026 YoY
for i in range(4):
    formula = f'=IF({quarterly_cols_2026[i]}{row_refs["seg_total"]}>0,{quarterly_cols_2026[i]}{row_refs["seg_total"]}/{quarterly_cols_2025[i]}{row_refs["seg_total"]}-1,"")'
    ws.cell(row=row, column=26+i, value=formula).number_format = pct_format
row += 1

row += 2  # Blank rows

# ==================== INCOME STATEMENT ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='INCOME STATEMENT - In Thousands USD').font = header_font
for i, yr in enumerate(annual_years_hist):
    ws.cell(row=row, column=3+i, value=yr).font = header_font
for i, yr in enumerate(annual_years_fcast):
    ws.cell(row=row, column=7+i, value=yr).font = header_font
for i, qtr in enumerate(quarterly_periods):
    ws.cell(row=row, column=18+i, value=qtr).font = header_font
row += 1

# Dates row
ws.cell(row=row, column=2, value='Period Ending')
for i, dt in enumerate(annual_dates_hist):
    ws.cell(row=row, column=3+i, value=dt)
for i, dt in enumerate(annual_dates_fcast):
    ws.cell(row=row, column=7+i, value=dt)
for i, dt in enumerate(quarterly_dates):
    ws.cell(row=row, column=18+i, value=dt)
row += 1

# Revenue
row_refs['revenue'] = row
ws.cell(row=row, column=2, value='Revenue').font = Font(bold=True)
for i, val in enumerate(annual_data['revenue']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
# Forecast annual - leave blank for now
for i, val in enumerate(quarterly_data_2024['revenue']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2025['revenue']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
# Q1-Q4 2026 left blank for projections
row += 1

# Cost of Revenue
row_refs['cost_of_revenue'] = row
ws.cell(row=row, column=2, value='- Cost of Revenue')
for i, val in enumerate(annual_data['cost_of_revenue']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2024['cost_of_revenue']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2025['cost_of_revenue']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
row += 1

# Cost of Revenue % of Sales (driver row with light gray shading)
row_refs['cost_pct_sales'] = row
ws.cell(row=row, column=2, value='  % of Sales')
for i, col in enumerate(annual_cols_hist):
    cell = ws.cell(row=row, column=3+i, value=f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["cost_of_revenue"]}/{col}{row_refs["revenue"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
for i, col in enumerate(annual_cols_fcast):
    cell = ws.cell(row=row, column=7+i, value=f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["cost_of_revenue"]}/{col}{row_refs["revenue"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
for i, col in enumerate(quarterly_cols):
    cell = ws.cell(row=row, column=18+i, value=f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["cost_of_revenue"]}/{col}{row_refs["revenue"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
# Apply fill to label cell too
ws.cell(row=row, column=2).fill = driver_fill
row += 1

# Gross Profit (formula)
row_refs['gross_profit'] = row
ws.cell(row=row, column=2, value='Gross Profit').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    ws.cell(row=row, column=3+i, value=f'={col}{row_refs["revenue"]}-{col}{row_refs["cost_of_revenue"]}').number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    ws.cell(row=row, column=7+i, value=f'={col}{row_refs["revenue"]}-{col}{row_refs["cost_of_revenue"]}').number_format = number_format
for i, col in enumerate(quarterly_cols):
    ws.cell(row=row, column=18+i, value=f'={col}{row_refs["revenue"]}-{col}{row_refs["cost_of_revenue"]}').number_format = number_format
row += 1

# SG&A
row_refs['sga'] = row
ws.cell(row=row, column=2, value='- SG&A Expenses')
for i, val in enumerate(annual_data['sga']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2024['sga']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2025['sga']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
row += 1

# SG&A % of Sales (driver row with light gray shading)
row_refs['sga_pct_sales'] = row
ws.cell(row=row, column=2, value='  % of Sales')
for i, col in enumerate(annual_cols_hist):
    cell = ws.cell(row=row, column=3+i, value=f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["sga"]}/{col}{row_refs["revenue"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
for i, col in enumerate(annual_cols_fcast):
    cell = ws.cell(row=row, column=7+i, value=f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["sga"]}/{col}{row_refs["revenue"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
for i, col in enumerate(quarterly_cols):
    cell = ws.cell(row=row, column=18+i, value=f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["sga"]}/{col}{row_refs["revenue"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
# Apply fill to label cell too
ws.cell(row=row, column=2).fill = driver_fill
row += 1

# D&A
row_refs['da'] = row
ws.cell(row=row, column=2, value='- Depreciation & Amortization')
for i, val in enumerate(annual_data['da']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2024['da']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2025['da']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
row += 1

# Goodwill Impairment
row_refs['goodwill_impairment'] = row
ws.cell(row=row, column=2, value='- Goodwill Impairment')
for i, val in enumerate(annual_data['goodwill_impairment']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2024['goodwill_impairment']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2025['goodwill_impairment']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
row += 1

# Operating Income (formula)
row_refs['operating_income'] = row
ws.cell(row=row, column=2, value='Operating Income (Loss)').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["gross_profit"]}-{col}{row_refs["sga"]}-{col}{row_refs["da"]}-{col}{row_refs["goodwill_impairment"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["gross_profit"]}-{col}{row_refs["sga"]}-{col}{row_refs["da"]}-{col}{row_refs["goodwill_impairment"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
for i, col in enumerate(quarterly_cols):
    formula = f'={col}{row_refs["gross_profit"]}-{col}{row_refs["sga"]}-{col}{row_refs["da"]}-{col}{row_refs["goodwill_impairment"]}'
    ws.cell(row=row, column=18+i, value=formula).number_format = number_format
row += 1

# Interest Expense
row_refs['interest_expense'] = row
ws.cell(row=row, column=2, value='- Interest Expense, net')
for i, val in enumerate(annual_data['interest_expense']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2024['interest_expense']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2025['interest_expense']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
row += 1

# Pretax Income (formula)
row_refs['pretax_income'] = row
ws.cell(row=row, column=2, value='Pretax Income (Loss)').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["operating_income"]}-{col}{row_refs["interest_expense"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["operating_income"]}-{col}{row_refs["interest_expense"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
for i, col in enumerate(quarterly_cols):
    formula = f'={col}{row_refs["operating_income"]}-{col}{row_refs["interest_expense"]}'
    ws.cell(row=row, column=18+i, value=formula).number_format = number_format
row += 1

# Tax Expense
row_refs['tax_expense'] = row
ws.cell(row=row, column=2, value='- Income Tax Expense (Benefit)')
for i, val in enumerate(annual_data['tax_expense']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2024['tax_expense']):
    ws.cell(row=row, column=18+i, value=val).number_format = number_format
for i, val in enumerate(quarterly_data_2025['tax_expense']):
    ws.cell(row=row, column=22+i, value=val).number_format = number_format
row += 1

# Net Income (formula)
row_refs['net_income'] = row
ws.cell(row=row, column=2, value='Net Income (Loss)').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["pretax_income"]}-{col}{row_refs["tax_expense"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["pretax_income"]}-{col}{row_refs["tax_expense"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
for i, col in enumerate(quarterly_cols):
    formula = f'={col}{row_refs["pretax_income"]}-{col}{row_refs["tax_expense"]}'
    ws.cell(row=row, column=18+i, value=formula).number_format = number_format
row += 1

row += 1  # Blank row

# Reference Items section
ws.cell(row=row, column=2, value='Reference Items').font = header_font
row += 1

# EBITDA (formula)
row_refs['ebitda'] = row
ws.cell(row=row, column=2, value='EBITDA')
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["operating_income"]}+{col}{row_refs["da"]}+{col}{row_refs["goodwill_impairment"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["operating_income"]}+{col}{row_refs["da"]}+{col}{row_refs["goodwill_impairment"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
for i, col in enumerate(quarterly_cols):
    formula = f'={col}{row_refs["operating_income"]}+{col}{row_refs["da"]}+{col}{row_refs["goodwill_impairment"]}'
    ws.cell(row=row, column=18+i, value=formula).number_format = number_format
row += 1

# Gross Margin (formula)
row_refs['gross_margin'] = row
ws.cell(row=row, column=2, value='Gross Margin')
for i, col in enumerate(annual_cols_hist):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["gross_profit"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["gross_profit"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
for i, col in enumerate(quarterly_cols):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["gross_profit"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=18+i, value=formula).number_format = pct_format
row += 1

# Operating Margin (formula)
row_refs['operating_margin'] = row
ws.cell(row=row, column=2, value='Operating Margin')
for i, col in enumerate(annual_cols_hist):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["operating_income"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["operating_income"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
for i, col in enumerate(quarterly_cols):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["operating_income"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=18+i, value=formula).number_format = pct_format
row += 1

# Net Profit Margin (formula)
row_refs['net_margin'] = row
ws.cell(row=row, column=2, value='Net Profit Margin')
for i, col in enumerate(annual_cols_hist):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["net_income"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["net_income"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
for i, col in enumerate(quarterly_cols):
    formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs["net_income"]}/{col}{row_refs["revenue"]},"")'
    ws.cell(row=row, column=18+i, value=formula).number_format = pct_format
row += 1

row += 2  # Blank rows

# ==================== TAX SCHEDULE ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='TAX SCHEDULE').font = header_font
for i, yr in enumerate(annual_years_hist):
    ws.cell(row=row, column=3+i, value=yr).font = header_font
for i, yr in enumerate(annual_years_fcast):
    ws.cell(row=row, column=7+i, value=yr).font = header_font
for i, qtr in enumerate(quarterly_periods):
    ws.cell(row=row, column=18+i, value=qtr).font = header_font
row += 1

# Pretax Income (reference from IS)
row_refs['tax_pretax'] = row
ws.cell(row=row, column=2, value='Pretax Income (Loss)')
for i, col in enumerate(annual_cols_hist):
    ws.cell(row=row, column=3+i, value=f'={col}{row_refs["pretax_income"]}').number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    ws.cell(row=row, column=7+i, value=f'={col}{row_refs["pretax_income"]}').number_format = number_format
for i, col in enumerate(quarterly_cols):
    ws.cell(row=row, column=18+i, value=f'={col}{row_refs["pretax_income"]}').number_format = number_format
row += 1

# Tax Expense (reference from IS)
row_refs['tax_expense_sched'] = row
ws.cell(row=row, column=2, value='Income Tax Expense (Benefit)')
for i, col in enumerate(annual_cols_hist):
    ws.cell(row=row, column=3+i, value=f'={col}{row_refs["tax_expense"]}').number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    ws.cell(row=row, column=7+i, value=f'={col}{row_refs["tax_expense"]}').number_format = number_format
for i, col in enumerate(quarterly_cols):
    ws.cell(row=row, column=18+i, value=f'={col}{row_refs["tax_expense"]}').number_format = number_format
row += 1

# Effective Tax Rate (driver row with shading)
row_refs['tax_rate'] = row
ws.cell(row=row, column=2, value='Effective Tax Rate')
for i, col in enumerate(annual_cols_hist):
    cell = ws.cell(row=row, column=3+i, value=f'=IF({col}{row_refs["tax_pretax"]}<>0,{col}{row_refs["tax_expense_sched"]}/{col}{row_refs["tax_pretax"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
for i, col in enumerate(annual_cols_fcast):
    cell = ws.cell(row=row, column=7+i, value=f'=IF({col}{row_refs["tax_pretax"]}<>0,{col}{row_refs["tax_expense_sched"]}/{col}{row_refs["tax_pretax"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
for i, col in enumerate(quarterly_cols):
    cell = ws.cell(row=row, column=18+i, value=f'=IF({col}{row_refs["tax_pretax"]}<>0,{col}{row_refs["tax_expense_sched"]}/{col}{row_refs["tax_pretax"]},"")')
    cell.number_format = pct_format
    cell.fill = driver_fill
ws.cell(row=row, column=2).fill = driver_fill
row += 1

row += 2

# ==================== BALANCE SHEET (ANNUAL ONLY) ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='BALANCE SHEET - In Thousands of USD').font = header_font
for i, yr in enumerate(annual_years_hist):
    ws.cell(row=row, column=3+i, value=yr).font = header_font
for i, yr in enumerate(annual_years_fcast):
    ws.cell(row=row, column=7+i, value=yr).font = header_font
# NO QUARTERLY DATA FOR BALANCE SHEET
row += 1

ws.cell(row=row, column=2, value='Period Ending')
for i, dt in enumerate(annual_dates_hist):
    ws.cell(row=row, column=3+i, value=dt)
for i, dt in enumerate(annual_dates_fcast):
    ws.cell(row=row, column=7+i, value=dt)
row += 1

ws.cell(row=row, column=2, value='ASSETS').font = header_font
row += 1

# Cash
row_refs['cash'] = row
ws.cell(row=row, column=2, value='+ Cash and Cash Equivalents')
for i, val in enumerate(annual_data['cash']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Accounts Receivable
row_refs['accounts_receivable'] = row
ws.cell(row=row, column=2, value='+ Accounts Receivable, net')
for i, val in enumerate(annual_data['accounts_receivable']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Subcontractor Receivables
row_refs['subcontractor_recv'] = row
ws.cell(row=row, column=2, value='+ Subcontractor Receivables')
for i, val in enumerate(annual_data['subcontractor_recv']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Prepaid & Other
row_refs['prepaid_other'] = row
ws.cell(row=row, column=2, value='+ Prepaid & Other Current')
for i, val in enumerate(annual_data['prepaid_other']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Total Current Assets (formula)
row_refs['total_current_assets'] = row
ws.cell(row=row, column=2, value='Total Current Assets').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["cash"]}+{col}{row_refs["accounts_receivable"]}+{col}{row_refs["subcontractor_recv"]}+{col}{row_refs["prepaid_other"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["cash"]}+{col}{row_refs["accounts_receivable"]}+{col}{row_refs["subcontractor_recv"]}+{col}{row_refs["prepaid_other"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

# Restricted Cash
row_refs['restricted_cash'] = row
ws.cell(row=row, column=2, value='+ Restricted Cash & Investments')
for i, val in enumerate(annual_data['restricted_cash']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Fixed Assets
row_refs['fixed_assets'] = row
ws.cell(row=row, column=2, value='+ Fixed Assets, net')
for i, val in enumerate(annual_data['fixed_assets']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Other Assets
row_refs['other_assets'] = row
ws.cell(row=row, column=2, value='+ Other Assets')
for i, val in enumerate(annual_data['other_assets']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Deferred Tax Assets
row_refs['deferred_tax_asset'] = row
ws.cell(row=row, column=2, value='+ Deferred Tax Assets')
for i, val in enumerate(annual_data['deferred_tax_asset']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Goodwill
row_refs['goodwill'] = row
ws.cell(row=row, column=2, value='+ Goodwill')
for i, val in enumerate(annual_data['goodwill']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Intangibles
row_refs['intangibles'] = row
ws.cell(row=row, column=2, value='+ Intangible Assets, net')
for i, val in enumerate(annual_data['intangibles']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Total Assets (formula)
row_refs['total_assets'] = row
ws.cell(row=row, column=2, value='Total Assets').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["total_current_assets"]}+{col}{row_refs["restricted_cash"]}+{col}{row_refs["fixed_assets"]}+{col}{row_refs["other_assets"]}+{col}{row_refs["deferred_tax_asset"]}+{col}{row_refs["goodwill"]}+{col}{row_refs["intangibles"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["total_current_assets"]}+{col}{row_refs["restricted_cash"]}+{col}{row_refs["fixed_assets"]}+{col}{row_refs["other_assets"]}+{col}{row_refs["deferred_tax_asset"]}+{col}{row_refs["goodwill"]}+{col}{row_refs["intangibles"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

row += 1
ws.cell(row=row, column=2, value="LIABILITIES & SHAREHOLDERS' EQUITY").font = header_font
row += 1

# AP & Accrued
row_refs['ap_accrued'] = row
ws.cell(row=row, column=2, value='+ Accounts Payable & Accrued')
for i, val in enumerate(annual_data['ap_accrued']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Accrued Compensation
row_refs['accrued_compensation'] = row
ws.cell(row=row, column=2, value='+ Accrued Compensation & Benefits')
for i, val in enumerate(annual_data['accrued_compensation']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Other Current Liabilities
row_refs['other_current_liab'] = row
ws.cell(row=row, column=2, value='+ Other Current Liabilities')
for i, val in enumerate(annual_data['other_current_liab']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Total Current Liabilities (formula)
row_refs['total_current_liab'] = row
ws.cell(row=row, column=2, value='Total Current Liabilities').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["ap_accrued"]}+{col}{row_refs["accrued_compensation"]}+{col}{row_refs["other_current_liab"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["ap_accrued"]}+{col}{row_refs["accrued_compensation"]}+{col}{row_refs["other_current_liab"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

# Revolving Credit
row_refs['revolving_credit'] = row
ws.cell(row=row, column=2, value='+ Revolving Credit Facility')
for i, val in enumerate(annual_data['revolving_credit']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Notes Payable
row_refs['notes_payable'] = row
ws.cell(row=row, column=2, value='+ Notes Payable, net')
for i, val in enumerate(annual_data['notes_payable']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Deferred Tax Liabilities
row_refs['deferred_tax_liab'] = row
ws.cell(row=row, column=2, value='+ Deferred Income Taxes')
for i, val in enumerate(annual_data['deferred_tax_liab']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Other LT Liabilities
row_refs['other_lt_liab'] = row
ws.cell(row=row, column=2, value='+ Other Long-term Liabilities')
for i, val in enumerate(annual_data['other_lt_liab']):
    ws.cell(row=row, column=3+i, value=val).number_format = number_format
row += 1

# Total Liabilities (formula)
row_refs['total_liabilities'] = row
ws.cell(row=row, column=2, value='Total Liabilities').font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["total_current_liab"]}+{col}{row_refs["revolving_credit"]}+{col}{row_refs["notes_payable"]}+{col}{row_refs["deferred_tax_liab"]}+{col}{row_refs["other_lt_liab"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["total_current_liab"]}+{col}{row_refs["revolving_credit"]}+{col}{row_refs["notes_payable"]}+{col}{row_refs["deferred_tax_liab"]}+{col}{row_refs["other_lt_liab"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

# Stockholders' Equity
row_refs['stockholders_equity'] = row
ws.cell(row=row, column=2, value="Stockholders' Equity").font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["total_assets"]}-{col}{row_refs["total_liabilities"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["total_assets"]}-{col}{row_refs["total_liabilities"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

# Total Liabilities & Equity
row_refs['total_liab_equity'] = row
ws.cell(row=row, column=2, value="Total Liabilities & Equity").font = Font(bold=True)
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["total_liabilities"]}+{col}{row_refs["stockholders_equity"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["total_liabilities"]}+{col}{row_refs["stockholders_equity"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

row += 1
# Check A=L+E
row_refs['balance_check'] = row
ws.cell(row=row, column=2, value='Check A=L+E')
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["total_assets"]}-{col}{row_refs["total_liab_equity"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["total_assets"]}-{col}{row_refs["total_liab_equity"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

row += 2

# ==================== INTEREST SCHEDULE ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='INTEREST EXPENSE AND DEBT SCHEDULE').font = header_font
for i, yr in enumerate(annual_years_hist):
    ws.cell(row=row, column=3+i, value=yr).font = header_font
for i, yr in enumerate(annual_years_fcast):
    ws.cell(row=row, column=7+i, value=yr).font = header_font
row += 1

row_refs['int_sched_expense'] = row
ws.cell(row=row, column=2, value='Interest Expense')
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["interest_expense"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["interest_expense"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

row_refs['total_debt'] = row
ws.cell(row=row, column=2, value='Total Debt')
for i, col in enumerate(annual_cols_hist):
    formula = f'={col}{row_refs["revolving_credit"]}+{col}{row_refs["notes_payable"]}'
    ws.cell(row=row, column=3+i, value=formula).number_format = number_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["revolving_credit"]}+{col}{row_refs["notes_payable"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

row_refs['implied_rate'] = row
ws.cell(row=row, column=2, value='Implied Interest Rate')
for i, col in enumerate(annual_cols_hist):
    formula = f'=IF({col}{row_refs["total_debt"]}>0,{col}{row_refs["int_sched_expense"]}/{col}{row_refs["total_debt"]},0)'
    ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
for i, col in enumerate(annual_cols_fcast):
    formula = f'=IF({col}{row_refs["total_debt"]}>0,{col}{row_refs["int_sched_expense"]}/{col}{row_refs["total_debt"]},0)'
    ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
row += 1

row_refs['debt_to_ebitda'] = row
ws.cell(row=row, column=2, value='Debt / EBITDA')
for i, col in enumerate(annual_cols_hist):
    formula = f'=IF({col}{row_refs["ebitda"]}>0,{col}{row_refs["total_debt"]}/{col}{row_refs["ebitda"]},0)'
    ws.cell(row=row, column=3+i, value=formula).number_format = '0.0x'
for i, col in enumerate(annual_cols_fcast):
    formula = f'=IF({col}{row_refs["ebitda"]}>0,{col}{row_refs["total_debt"]}/{col}{row_refs["ebitda"]},0)'
    ws.cell(row=row, column=7+i, value=formula).number_format = '0.0x'
row += 1

row += 2

# ==================== PPE AND CAPEX SCHEDULE ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='PPE AND CAPEX SCHEDULE').font = header_font
for i, yr in enumerate(['FY 2022', 'FY 2023', 'FY 2024']):
    ws.cell(row=row, column=4+i, value=yr).font = header_font
row += 1

row_refs['capex'] = row
ws.cell(row=row, column=2, value='CapEx (implied)')
ws.cell(row=row, column=4, value=f'=D{row_refs["fixed_assets"]}-C{row_refs["fixed_assets"]}+D{row_refs["da"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["fixed_assets"]}-D{row_refs["fixed_assets"]}+E{row_refs["da"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["fixed_assets"]}-E{row_refs["fixed_assets"]}+F{row_refs["da"]}').number_format = number_format
row += 1

row_refs['ppe_revenue'] = row
ws.cell(row=row, column=2, value='Revenue')
ws.cell(row=row, column=4, value=f'=D{row_refs["revenue"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["revenue"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["revenue"]}').number_format = number_format
row += 1

row_refs['capex_pct'] = row
ws.cell(row=row, column=2, value='CapEx as % of Revenue')
cell = ws.cell(row=row, column=4, value=f'=D{row_refs["capex"]}/D{row_refs["ppe_revenue"]}')
cell.number_format = pct_format
cell.fill = driver_fill
cell = ws.cell(row=row, column=5, value=f'=E{row_refs["capex"]}/E{row_refs["ppe_revenue"]}')
cell.number_format = pct_format
cell.fill = driver_fill
cell = ws.cell(row=row, column=6, value=f'=F{row_refs["capex"]}/F{row_refs["ppe_revenue"]}')
cell.number_format = pct_format
cell.fill = driver_fill
ws.cell(row=row, column=2).fill = driver_fill
row += 1

row += 1

row_refs['ppe_da'] = row
ws.cell(row=row, column=2, value='Depreciation & Amortization')
ws.cell(row=row, column=4, value=f'=D{row_refs["da"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["da"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["da"]}').number_format = number_format
row += 1

row_refs['da_pct'] = row
ws.cell(row=row, column=2, value='D&A as % of Revenue')
cell = ws.cell(row=row, column=4, value=f'=D{row_refs["ppe_da"]}/D{row_refs["ppe_revenue"]}')
cell.number_format = pct_format
cell.fill = driver_fill
cell = ws.cell(row=row, column=5, value=f'=E{row_refs["ppe_da"]}/E{row_refs["ppe_revenue"]}')
cell.number_format = pct_format
cell.fill = driver_fill
cell = ws.cell(row=row, column=6, value=f'=F{row_refs["ppe_da"]}/F{row_refs["ppe_revenue"]}')
cell.number_format = pct_format
cell.fill = driver_fill
ws.cell(row=row, column=2).fill = driver_fill
row += 1

row += 2

# ==================== COMMON SIZE BALANCE SHEET ====================
ws.cell(row=row, column=2, value='COMMON SIZE BALANCE SHEET').font = header_font
row += 1
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='Balance Sheet Items as % of Revenue').font = header_font
for i, yr in enumerate(annual_years_hist):
    ws.cell(row=row, column=3+i, value=yr).font = header_font
for i, yr in enumerate(annual_years_fcast):
    ws.cell(row=row, column=7+i, value=yr).font = header_font
row += 1

ws.cell(row=row, column=2, value='ASSETS').font = header_font
row += 1

common_size_assets = [
    ('+ Cash and Cash Equivalents', 'cash'),
    ('+ Accounts Receivable, net', 'accounts_receivable'),
    ('+ Subcontractor Receivables', 'subcontractor_recv'),
    ('+ Prepaid & Other Current', 'prepaid_other'),
    ('Total Current Assets', 'total_current_assets'),
    ('+ Restricted Cash & Investments', 'restricted_cash'),
    ('+ Fixed Assets, net', 'fixed_assets'),
    ('+ Other Assets', 'other_assets'),
    ('+ Deferred Tax Assets', 'deferred_tax_asset'),
    ('+ Goodwill', 'goodwill'),
    ('+ Intangible Assets, net', 'intangibles'),
    ('Total Assets', 'total_assets'),
]

for label, key in common_size_assets:
    bold = 'Total' in label
    ws.cell(row=row, column=2, value=label)
    if bold:
        ws.cell(row=row, column=2).font = Font(bold=True)
    for i, col in enumerate(annual_cols_hist):
        formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs[key]}/{col}{row_refs["revenue"]},"")'
        ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
    for i, col in enumerate(annual_cols_fcast):
        formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs[key]}/{col}{row_refs["revenue"]},"")'
        ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
    row += 1

row += 1
ws.cell(row=row, column=2, value="LIABILITIES & EQUITY").font = header_font
row += 1

common_size_liab = [
    ('+ Accounts Payable & Accrued', 'ap_accrued'),
    ('+ Accrued Compensation & Benefits', 'accrued_compensation'),
    ('+ Other Current Liabilities', 'other_current_liab'),
    ('Total Current Liabilities', 'total_current_liab'),
    ('+ Revolving Credit Facility', 'revolving_credit'),
    ('+ Notes Payable, net', 'notes_payable'),
    ('+ Deferred Income Taxes', 'deferred_tax_liab'),
    ('+ Other Long-term Liabilities', 'other_lt_liab'),
    ('Total Liabilities', 'total_liabilities'),
    ("Stockholders' Equity", 'stockholders_equity'),
    ("Total Liabilities & Equity", 'total_liab_equity'),
]

for label, key in common_size_liab:
    bold = 'Total' in label or 'Equity' in label
    ws.cell(row=row, column=2, value=label)
    if bold:
        ws.cell(row=row, column=2).font = Font(bold=True)
    for i, col in enumerate(annual_cols_hist):
        formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs[key]}/{col}{row_refs["revenue"]},"")'
        ws.cell(row=row, column=3+i, value=formula).number_format = pct_format
    for i, col in enumerate(annual_cols_fcast):
        formula = f'=IF({col}{row_refs["revenue"]}>0,{col}{row_refs[key]}/{col}{row_refs["revenue"]},"")'
        ws.cell(row=row, column=7+i, value=formula).number_format = pct_format
    row += 1

row += 2

# ==================== EQUITY SCHEDULE ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='EQUITY SCHEDULE').font = header_font
for i, yr in enumerate(['FY 2022', 'FY 2023', 'FY 2024']):
    ws.cell(row=row, column=4+i, value=yr).font = header_font
row += 1

row_refs['equity_bop'] = row
ws.cell(row=row, column=2, value='Beginning of Period').font = Font(bold=True)
ws.cell(row=row, column=4, value=f'=C{row_refs["stockholders_equity"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=D{row_refs["stockholders_equity"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=E{row_refs["stockholders_equity"]}').number_format = number_format
row += 1

row_refs['equity_ni'] = row
ws.cell(row=row, column=2, value='+ Net Income')
ws.cell(row=row, column=4, value=f'=D{row_refs["net_income"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["net_income"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["net_income"]}').number_format = number_format
row += 1

row_refs['equity_other'] = row
ws.cell(row=row, column=2, value='+ Other Changes (Buybacks/Dividends/OCI)')
ws.cell(row=row, column=4, value=f'=D{row_refs["stockholders_equity"]}-D{row_refs["equity_bop"]}-D{row_refs["equity_ni"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["stockholders_equity"]}-E{row_refs["equity_bop"]}-E{row_refs["equity_ni"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["stockholders_equity"]}-F{row_refs["equity_bop"]}-F{row_refs["equity_ni"]}').number_format = number_format
row += 1

row_refs['equity_eop'] = row
ws.cell(row=row, column=2, value='End of Period').font = Font(bold=True)
ws.cell(row=row, column=4, value=f'=D{row_refs["stockholders_equity"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["stockholders_equity"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["stockholders_equity"]}').number_format = number_format
row += 1

row += 2

# ==================== STATEMENT OF CASH FLOWS ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='STATEMENT OF CASH FLOWS (Derived)').font = header_font
for i, yr in enumerate(['FY 2022', 'FY 2023', 'FY 2024']):
    ws.cell(row=row, column=4+i, value=yr).font = header_font
row += 1

ws.cell(row=row, column=2, value='Operating Activities:').font = header_font
row += 1

row_refs['scf_ni'] = row
ws.cell(row=row, column=2, value='Net Income')
ws.cell(row=row, column=4, value=f'=D{row_refs["net_income"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["net_income"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["net_income"]}').number_format = number_format
row += 1

row_refs['scf_da'] = row
ws.cell(row=row, column=2, value='+ Depreciation & Amortization')
ws.cell(row=row, column=4, value=f'=D{row_refs["da"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["da"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["da"]}').number_format = number_format
row += 1

row_refs['scf_impairment'] = row
ws.cell(row=row, column=2, value='+ Goodwill Impairment')
ws.cell(row=row, column=4, value=f'=D{row_refs["goodwill_impairment"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["goodwill_impairment"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["goodwill_impairment"]}').number_format = number_format
row += 1

row_refs['scf_ar'] = row
ws.cell(row=row, column=2, value='change in + Accounts Receivable')
ws.cell(row=row, column=4, value=f'=-(D{row_refs["accounts_receivable"]}-C{row_refs["accounts_receivable"]})').number_format = number_format
ws.cell(row=row, column=5, value=f'=-(E{row_refs["accounts_receivable"]}-D{row_refs["accounts_receivable"]})').number_format = number_format
ws.cell(row=row, column=6, value=f'=-(F{row_refs["accounts_receivable"]}-E{row_refs["accounts_receivable"]})').number_format = number_format
row += 1

row_refs['scf_sub'] = row
ws.cell(row=row, column=2, value='change in + Subcontractor Receivables')
ws.cell(row=row, column=4, value=f'=-(D{row_refs["subcontractor_recv"]}-C{row_refs["subcontractor_recv"]})').number_format = number_format
ws.cell(row=row, column=5, value=f'=-(E{row_refs["subcontractor_recv"]}-D{row_refs["subcontractor_recv"]})').number_format = number_format
ws.cell(row=row, column=6, value=f'=-(F{row_refs["subcontractor_recv"]}-E{row_refs["subcontractor_recv"]})').number_format = number_format
row += 1

row_refs['scf_prepaid'] = row
ws.cell(row=row, column=2, value='change in + Prepaid & Other')
ws.cell(row=row, column=4, value=f'=-(D{row_refs["prepaid_other"]}-C{row_refs["prepaid_other"]})').number_format = number_format
ws.cell(row=row, column=5, value=f'=-(E{row_refs["prepaid_other"]}-D{row_refs["prepaid_other"]})').number_format = number_format
ws.cell(row=row, column=6, value=f'=-(F{row_refs["prepaid_other"]}-E{row_refs["prepaid_other"]})').number_format = number_format
row += 1

row_refs['scf_ap'] = row
ws.cell(row=row, column=2, value='change in + Payables & Accruals')
ws.cell(row=row, column=4, value=f'=D{row_refs["ap_accrued"]}-C{row_refs["ap_accrued"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["ap_accrued"]}-D{row_refs["ap_accrued"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["ap_accrued"]}-E{row_refs["ap_accrued"]}').number_format = number_format
row += 1

row_refs['scf_comp'] = row
ws.cell(row=row, column=2, value='change in + Accrued Compensation')
ws.cell(row=row, column=4, value=f'=D{row_refs["accrued_compensation"]}-C{row_refs["accrued_compensation"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["accrued_compensation"]}-D{row_refs["accrued_compensation"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["accrued_compensation"]}-E{row_refs["accrued_compensation"]}').number_format = number_format
row += 1

row_refs['scf_other_cl'] = row
ws.cell(row=row, column=2, value='change in + Other ST Liabilities')
ws.cell(row=row, column=4, value=f'=D{row_refs["other_current_liab"]}-C{row_refs["other_current_liab"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["other_current_liab"]}-D{row_refs["other_current_liab"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["other_current_liab"]}-E{row_refs["other_current_liab"]}').number_format = number_format
row += 1

row_refs['cfo'] = row
ws.cell(row=row, column=2, value='Cash Flow From Operations').font = Font(bold=True)
ws.cell(row=row, column=4, value=f'=D{row_refs["scf_ni"]}+D{row_refs["scf_da"]}+D{row_refs["scf_impairment"]}+D{row_refs["scf_ar"]}+D{row_refs["scf_sub"]}+D{row_refs["scf_prepaid"]}+D{row_refs["scf_ap"]}+D{row_refs["scf_comp"]}+D{row_refs["scf_other_cl"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["scf_ni"]}+E{row_refs["scf_da"]}+E{row_refs["scf_impairment"]}+E{row_refs["scf_ar"]}+E{row_refs["scf_sub"]}+E{row_refs["scf_prepaid"]}+E{row_refs["scf_ap"]}+E{row_refs["scf_comp"]}+E{row_refs["scf_other_cl"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["scf_ni"]}+F{row_refs["scf_da"]}+F{row_refs["scf_impairment"]}+F{row_refs["scf_ar"]}+F{row_refs["scf_sub"]}+F{row_refs["scf_prepaid"]}+F{row_refs["scf_ap"]}+F{row_refs["scf_comp"]}+F{row_refs["scf_other_cl"]}').number_format = number_format
row += 1

row += 1
ws.cell(row=row, column=2, value='Investing Activities:').font = header_font
row += 1

row_refs['scf_ppe'] = row
ws.cell(row=row, column=2, value='change in + Property, Plant & Equip')
ws.cell(row=row, column=4, value=f'=-((D{row_refs["fixed_assets"]}-C{row_refs["fixed_assets"]})+D{row_refs["da"]})').number_format = number_format
ws.cell(row=row, column=5, value=f'=-((E{row_refs["fixed_assets"]}-D{row_refs["fixed_assets"]})+E{row_refs["da"]})').number_format = number_format
ws.cell(row=row, column=6, value=f'=-((F{row_refs["fixed_assets"]}-E{row_refs["fixed_assets"]})+F{row_refs["da"]})').number_format = number_format
row += 1

row_refs['scf_other_lt_assets'] = row
ws.cell(row=row, column=2, value='change in + Other LT Assets')
ws.cell(row=row, column=4, value=f'=-((D{row_refs["restricted_cash"]}-C{row_refs["restricted_cash"]})+(D{row_refs["other_assets"]}-C{row_refs["other_assets"]})+(D{row_refs["deferred_tax_asset"]}-C{row_refs["deferred_tax_asset"]}))').number_format = number_format
ws.cell(row=row, column=5, value=f'=-((E{row_refs["restricted_cash"]}-D{row_refs["restricted_cash"]})+(E{row_refs["other_assets"]}-D{row_refs["other_assets"]})+(E{row_refs["deferred_tax_asset"]}-D{row_refs["deferred_tax_asset"]}))').number_format = number_format
ws.cell(row=row, column=6, value=f'=-((F{row_refs["restricted_cash"]}-E{row_refs["restricted_cash"]})+(F{row_refs["other_assets"]}-E{row_refs["other_assets"]})+(F{row_refs["deferred_tax_asset"]}-E{row_refs["deferred_tax_asset"]}))').number_format = number_format
row += 1

row_refs['scf_goodwill'] = row
ws.cell(row=row, column=2, value='change in + Goodwill (Acquisitions)')
ws.cell(row=row, column=4, value=f'=-((D{row_refs["goodwill"]}-C{row_refs["goodwill"]})+D{row_refs["goodwill_impairment"]})').number_format = number_format
ws.cell(row=row, column=5, value=f'=-((E{row_refs["goodwill"]}-D{row_refs["goodwill"]})+E{row_refs["goodwill_impairment"]})').number_format = number_format
ws.cell(row=row, column=6, value=f'=-((F{row_refs["goodwill"]}-E{row_refs["goodwill"]})+F{row_refs["goodwill_impairment"]})').number_format = number_format
row += 1

row_refs['scf_intangibles'] = row
ws.cell(row=row, column=2, value='change in + Intangibles')
ws.cell(row=row, column=4, value=f'=-(D{row_refs["intangibles"]}-C{row_refs["intangibles"]})').number_format = number_format
ws.cell(row=row, column=5, value=f'=-(E{row_refs["intangibles"]}-D{row_refs["intangibles"]})').number_format = number_format
ws.cell(row=row, column=6, value=f'=-(F{row_refs["intangibles"]}-E{row_refs["intangibles"]})').number_format = number_format
row += 1

row_refs['cfi'] = row
ws.cell(row=row, column=2, value='Cash Flow From Investing').font = Font(bold=True)
ws.cell(row=row, column=4, value=f'=D{row_refs["scf_ppe"]}+D{row_refs["scf_other_lt_assets"]}+D{row_refs["scf_goodwill"]}+D{row_refs["scf_intangibles"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["scf_ppe"]}+E{row_refs["scf_other_lt_assets"]}+E{row_refs["scf_goodwill"]}+E{row_refs["scf_intangibles"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["scf_ppe"]}+F{row_refs["scf_other_lt_assets"]}+F{row_refs["scf_goodwill"]}+F{row_refs["scf_intangibles"]}').number_format = number_format
row += 1

row += 1
ws.cell(row=row, column=2, value='Financing Activities:').font = header_font
row += 1

row_refs['scf_lt_debt'] = row
ws.cell(row=row, column=2, value='change in + LT Debt')
ws.cell(row=row, column=4, value=f'=(D{row_refs["revolving_credit"]}-C{row_refs["revolving_credit"]})+(D{row_refs["notes_payable"]}-C{row_refs["notes_payable"]})').number_format = number_format
ws.cell(row=row, column=5, value=f'=(E{row_refs["revolving_credit"]}-D{row_refs["revolving_credit"]})+(E{row_refs["notes_payable"]}-D{row_refs["notes_payable"]})').number_format = number_format
ws.cell(row=row, column=6, value=f'=(F{row_refs["revolving_credit"]}-E{row_refs["revolving_credit"]})+(F{row_refs["notes_payable"]}-E{row_refs["notes_payable"]})').number_format = number_format
row += 1

row_refs['scf_other_lt_liab'] = row
ws.cell(row=row, column=2, value='change in + Other LT Liabilities')
ws.cell(row=row, column=4, value=f'=(D{row_refs["deferred_tax_liab"]}-C{row_refs["deferred_tax_liab"]})+(D{row_refs["other_lt_liab"]}-C{row_refs["other_lt_liab"]})').number_format = number_format
ws.cell(row=row, column=5, value=f'=(E{row_refs["deferred_tax_liab"]}-D{row_refs["deferred_tax_liab"]})+(E{row_refs["other_lt_liab"]}-D{row_refs["other_lt_liab"]})').number_format = number_format
ws.cell(row=row, column=6, value=f'=(F{row_refs["deferred_tax_liab"]}-E{row_refs["deferred_tax_liab"]})+(F{row_refs["other_lt_liab"]}-E{row_refs["other_lt_liab"]})').number_format = number_format
row += 1

row_refs['scf_equity'] = row
ws.cell(row=row, column=2, value='change in Total Equity (ex NI)')
ws.cell(row=row, column=4, value=f'=D{row_refs["equity_other"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["equity_other"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["equity_other"]}').number_format = number_format
row += 1

row_refs['cff'] = row
ws.cell(row=row, column=2, value='Cash Flow From Financing').font = Font(bold=True)
ws.cell(row=row, column=4, value=f'=D{row_refs["scf_lt_debt"]}+D{row_refs["scf_other_lt_liab"]}+D{row_refs["scf_equity"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["scf_lt_debt"]}+E{row_refs["scf_other_lt_liab"]}+E{row_refs["scf_equity"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["scf_lt_debt"]}+F{row_refs["scf_other_lt_liab"]}+F{row_refs["scf_equity"]}').number_format = number_format
row += 1

row += 1

row_refs['total_cf'] = row
ws.cell(row=row, column=2, value='Total Cash Flow').font = Font(bold=True)
ws.cell(row=row, column=4, value=f'=D{row_refs["cfo"]}+D{row_refs["cfi"]}+D{row_refs["cff"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["cfo"]}+E{row_refs["cfi"]}+E{row_refs["cff"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["cfo"]}+F{row_refs["cfi"]}+F{row_refs["cff"]}').number_format = number_format
row += 1

row += 1

row_refs['actual_cf'] = row
ws.cell(row=row, column=2, value='Actual Change in Cash (B/S)')
ws.cell(row=row, column=4, value=f'=D{row_refs["cash"]}-C{row_refs["cash"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["cash"]}-D{row_refs["cash"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["cash"]}-E{row_refs["cash"]}').number_format = number_format
row += 1

row_refs['cf_diff'] = row
ws.cell(row=row, column=2, value='Difference (Check)')
ws.cell(row=row, column=4, value=f'=D{row_refs["total_cf"]}-D{row_refs["actual_cf"]}').number_format = number_format
ws.cell(row=row, column=5, value=f'=E{row_refs["total_cf"]}-E{row_refs["actual_cf"]}').number_format = number_format
ws.cell(row=row, column=6, value=f'=F{row_refs["total_cf"]}-F{row_refs["actual_cf"]}').number_format = number_format
row += 1

row += 2

# ==================== DCF VALUATION ====================
# Define input fill color (light blue for inputs)
input_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='DCF VALUATION - Annual (in Thousands USD)').font = header_font
for i, yr in enumerate(annual_years_fcast):
    ws.cell(row=row, column=7+i, value=yr).font = header_font
row += 1

# DCF Inputs Section
ws.cell(row=row, column=2, value='INPUTS').font = header_font
row += 1

row_refs['discount_rate'] = row
ws.cell(row=row, column=2, value='Discount Rate (WACC)')
cell = ws.cell(row=row, column=7, value=0.10)  # Default 10%
cell.number_format = pct_format
cell.fill = input_fill
ws.cell(row=row, column=2).fill = input_fill
row += 1

row_refs['terminal_growth'] = row
ws.cell(row=row, column=2, value='Terminal Growth Rate')
cell = ws.cell(row=row, column=7, value=0.02)  # Default 2%
cell.number_format = pct_format
cell.fill = input_fill
ws.cell(row=row, column=2).fill = input_fill
row += 1

row += 1

# Free Cash Flow Section
ws.cell(row=row, column=2, value='FREE CASH FLOW').font = header_font
row += 1

row_refs['dcf_ebitda'] = row
ws.cell(row=row, column=2, value='EBITDA')
for i, col in enumerate(annual_cols_fcast):
    ws.cell(row=row, column=7+i, value=f'={col}{row_refs["ebitda"]}').number_format = number_format
row += 1

row_refs['dcf_da'] = row
ws.cell(row=row, column=2, value='- D&A (add back already in EBITDA)')
for i, col in enumerate(annual_cols_fcast):
    ws.cell(row=row, column=7+i, value=0).number_format = number_format
row += 1

row_refs['dcf_capex'] = row
ws.cell(row=row, column=2, value='- CapEx (as % of Revenue)')
# Use FY2024 CapEx % as assumption for forecast years
for i, col in enumerate(annual_cols_fcast):
    # CapEx = Revenue * CapEx% (use F column ratio from PPE schedule)
    ws.cell(row=row, column=7+i, value=f'=-{col}{row_refs["revenue"]}*F{row_refs["capex_pct"]}').number_format = number_format
row += 1

row_refs['dcf_chg_nwc'] = row
ws.cell(row=row, column=2, value='- Change in NWC (assumed 0)')
for i, col in enumerate(annual_cols_fcast):
    ws.cell(row=row, column=7+i, value=0).number_format = number_format
row += 1

row_refs['dcf_taxes'] = row
ws.cell(row=row, column=2, value='- Cash Taxes')
for i, col in enumerate(annual_cols_fcast):
    ws.cell(row=row, column=7+i, value=f'=-{col}{row_refs["tax_expense"]}').number_format = number_format
row += 1

row_refs['fcf'] = row
ws.cell(row=row, column=2, value='Free Cash Flow').font = Font(bold=True)
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["dcf_ebitda"]}+{col}{row_refs["dcf_capex"]}+{col}{row_refs["dcf_chg_nwc"]}+{col}{row_refs["dcf_taxes"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

row += 1

# Present Value Section
ws.cell(row=row, column=2, value='PRESENT VALUE').font = header_font
row += 1

row_refs['discount_factor'] = row
ws.cell(row=row, column=2, value='Discount Factor')
for i in range(len(annual_cols_fcast)):
    year_num = i + 1
    formula = f'=1/(1+$G${row_refs["discount_rate"]})^{year_num}'
    ws.cell(row=row, column=7+i, value=formula).number_format = '0.000'
row += 1

row_refs['pv_fcf'] = row
ws.cell(row=row, column=2, value='PV of FCF')
for i, col in enumerate(annual_cols_fcast):
    formula = f'={col}{row_refs["fcf"]}*{col}{row_refs["discount_factor"]}'
    ws.cell(row=row, column=7+i, value=formula).number_format = number_format
row += 1

row += 1

# Terminal Value and Enterprise Value
ws.cell(row=row, column=2, value='VALUATION SUMMARY').font = header_font
row += 1

row_refs['sum_pv_fcf'] = row
ws.cell(row=row, column=2, value='Sum of PV of FCFs')
# Sum all forecast year PV FCFs
pv_sum_formula = '+'.join([f'{col}{row_refs["pv_fcf"]}' for col in annual_cols_fcast])
ws.cell(row=row, column=7, value=f'={pv_sum_formula}').number_format = number_format
row += 1

row_refs['terminal_fcf'] = row
ws.cell(row=row, column=2, value='Terminal Year FCF')
ws.cell(row=row, column=7, value=f'=L{row_refs["fcf"]}').number_format = number_format  # FY2030 FCF
row += 1

row_refs['terminal_value'] = row
ws.cell(row=row, column=2, value='Terminal Value (Gordon Growth)')
# TV = FCF * (1+g) / (r - g)
ws.cell(row=row, column=7, value=f'=G{row_refs["terminal_fcf"]}*(1+$G${row_refs["terminal_growth"]})/($G${row_refs["discount_rate"]}-$G${row_refs["terminal_growth"]})').number_format = number_format
row += 1

row_refs['pv_terminal'] = row
ws.cell(row=row, column=2, value='PV of Terminal Value')
# Discount TV back 6 years
ws.cell(row=row, column=7, value=f'=G{row_refs["terminal_value"]}*L{row_refs["discount_factor"]}').number_format = number_format
row += 1

row_refs['enterprise_value'] = row
ws.cell(row=row, column=2, value='Enterprise Value').font = Font(bold=True)
ws.cell(row=row, column=7, value=f'=G{row_refs["sum_pv_fcf"]}+G{row_refs["pv_terminal"]}').number_format = number_format
row += 1

row += 2

# ==================== MARKET DATA & UPSIDE/DOWNSIDE ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='MARKET DATA & VALUATION COMPARISON').font = header_font
row += 1

ws.cell(row=row, column=2, value='MARKET INPUTS (Update Manually)').font = header_font
row += 1

row_refs['shares_out'] = row
ws.cell(row=row, column=2, value='Shares Outstanding (millions)')
cell = ws.cell(row=row, column=7, value=37.5)  # Approx AMN shares
cell.number_format = '0.0'
cell.fill = input_fill
ws.cell(row=row, column=2).fill = input_fill
row += 1

row_refs['stock_price'] = row
ws.cell(row=row, column=2, value='Current Stock Price ($)')
cell = ws.cell(row=row, column=7, value=25.00)  # Placeholder
cell.number_format = '$#,##0.00'
cell.fill = input_fill
ws.cell(row=row, column=2).fill = input_fill
row += 1

row_refs['market_cap'] = row
ws.cell(row=row, column=2, value='Market Cap (thousands)')
ws.cell(row=row, column=7, value=f'=G{row_refs["shares_out"]}*G{row_refs["stock_price"]}*1000').number_format = number_format
row += 1

row_refs['total_debt_val'] = row
ws.cell(row=row, column=2, value='Total Debt (FY2024)')
ws.cell(row=row, column=7, value=f'=F{row_refs["total_debt"]}').number_format = number_format
row += 1

row_refs['cash_val'] = row
ws.cell(row=row, column=2, value='Cash (FY2024)')
ws.cell(row=row, column=7, value=f'=F{row_refs["cash"]}').number_format = number_format
row += 1

row_refs['net_debt'] = row
ws.cell(row=row, column=2, value='Net Debt').font = Font(bold=True)
ws.cell(row=row, column=7, value=f'=G{row_refs["total_debt_val"]}-G{row_refs["cash_val"]}').number_format = number_format
row += 1

row_refs['current_ev'] = row
ws.cell(row=row, column=2, value='Current Enterprise Value').font = Font(bold=True)
ws.cell(row=row, column=7, value=f'=G{row_refs["market_cap"]}+G{row_refs["net_debt"]}').number_format = number_format
row += 1

row += 1

ws.cell(row=row, column=2, value='VALUATION COMPARISON').font = header_font
row += 1

row_refs['dcf_ev'] = row
ws.cell(row=row, column=2, value='DCF Enterprise Value')
ws.cell(row=row, column=7, value=f'=G{row_refs["enterprise_value"]}').number_format = number_format
row += 1

row_refs['implied_equity'] = row
ws.cell(row=row, column=2, value='Implied Equity Value')
ws.cell(row=row, column=7, value=f'=G{row_refs["dcf_ev"]}-G{row_refs["net_debt"]}').number_format = number_format
row += 1

row_refs['implied_price'] = row
ws.cell(row=row, column=2, value='Implied Share Price ($)').font = Font(bold=True)
ws.cell(row=row, column=7, value=f'=G{row_refs["implied_equity"]}/G{row_refs["shares_out"]}/1000').number_format = '$#,##0.00'
row += 1

row_refs['upside_downside'] = row
ws.cell(row=row, column=2, value='Upside / (Downside) %').font = Font(bold=True)
cell = ws.cell(row=row, column=7, value=f'=G{row_refs["implied_price"]}/G{row_refs["stock_price"]}-1')
cell.number_format = pct_format
cell.font = Font(bold=True)
row += 1

row += 2

# ==================== DATA SOURCE REFERENCE ====================
ws.cell(row=row, column=1, value='x')
ws.cell(row=row, column=2, value='DATA SOURCES & FILING REFERENCES').font = header_font
row += 1

ws.cell(row=row, column=2, value='SEC EDGAR: https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=0001142750')
row += 1
ws.cell(row=row, column=2, value='Company: AMN Healthcare Services Inc (NYSE: AMN)')
row += 1
ws.cell(row=row, column=2, value='10-K (Annual): Revenue, all IS items, all BS items')
row += 1
ws.cell(row=row, column=2, value='10-Q (Quarterly): Quarterly IS data, segment revenue')
row += 1
ws.cell(row=row, column=2, value='Source File: Annual data from 10-K filings; Quarterly from 10-Q filings')
row += 1

row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 45
# Historical annual columns
for col_letter in ['C', 'D', 'E', 'F']:
    ws.column_dimensions[col_letter].width = 14
# Forecast annual columns
for col_letter in ['G', 'H', 'I', 'J', 'K', 'L']:
    ws.column_dimensions[col_letter].width = 14
# Blank separator columns
for col_letter in ['M', 'N', 'O', 'P', 'Q']:
    ws.column_dimensions[col_letter].width = 3
# Quarterly columns (2024, 2025, 2026)
for col_letter in ['R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
    ws.column_dimensions[col_letter].width = 12
# AA, AB, AC columns
ws.column_dimensions['AA'].width = 12
ws.column_dimensions['AB'].width = 12
ws.column_dimensions['AC'].width = 12

# Save workbook
wb.save('/home/user/ClaudeCodeGit/AMN_Healthcare_Model_v4.xlsx')
print('Excel model created successfully: AMN_Healthcare_Model_v4.xlsx')
print(f'Total rows: {row}')
